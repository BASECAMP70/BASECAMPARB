import os
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from dotenv import load_dotenv
import timesheet.storage as storage
from io import BytesIO
from datetime import date as dt_date, timedelta

import timesheet.pdf as pdf

load_dotenv()


def create_app(config=None):
    app = Flask(__name__)
    app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-timesheet")
    app.config["TEMPLATES_AUTO_RELOAD"] = True

    if config:
        app.config.update(config)

    @app.template_filter("money")
    def money_filter(value):
        return f"{float(value):,.2f}"

    # Inject settings into every template
    @app.context_processor
    def inject_settings():
        return {"settings": storage.load_settings()}

    def _migrate_tasks_to_global():
        """One-time migration: strip project_id/budgets from tasks, create project_task assignments."""
        tasks = storage.load_tasks()
        if not tasks or not any("project_id" in t for t in tasks):
            return
        pt = storage.load_project_tasks()
        existing_pt_keys = {(r["project_id"], r["task_id"]) for r in pt}
        changed = False
        for t in tasks:
            if "project_id" in t:
                pid = t.pop("project_id")
                key = (pid, t["id"])
                bh = t.pop("budget_hours", None)
                bd = t.pop("budget_dollars", None)
                if key not in existing_pt_keys:
                    pt.append({
                        "id": storage.new_id(),
                        "project_id": pid,
                        "task_id": t["id"],
                        "rate_override": None,
                        "budget_hours": bh,
                        "budget_dollars": bd,
                    })
                    existing_pt_keys.add(key)
                changed = True
        if changed:
            storage.save_tasks(tasks)
            storage.save_project_tasks(pt)

    def _migrate_expenses_to_client():
        """One-time migration: assign client_id to expenses that only have project_id."""
        items = storage.load_expenses()
        if not items or all("client_id" in x for x in items):
            return
        projects = storage.load_projects()
        project_map = {p["id"]: p for p in projects}
        changed = False
        for x in items:
            if "client_id" not in x:
                pid = x.get("project_id", "")
                p = project_map.get(pid, {})
                x["client_id"] = p.get("client_id") or ""
                changed = True
        if changed:
            storage.save_expenses(items)

    _migrate_tasks_to_global()

    @app.route("/")
    def dashboard():
        settings_data = storage.load_settings()
        gst_rate = settings_data.get("gst_rate", 0.05)
        all_invoices = storage.load_invoices()
        all_entries  = storage.load_time_entries()
        all_expenses = storage.load_expenses()
        projects     = storage.load_projects()
        project_map  = {p["id"]: p for p in projects}
        clients      = storage.load_clients()
        client_map   = {c["id"]: c for c in clients}

        # Outstanding invoices (not paid, excludes expense invoices)
        outstanding = [i for i in all_invoices
                       if i.get("status") not in ("paid",)
                       and i.get("invoice_type") != "expense"]
        outstanding_total = round(sum(i.get("total", 0) for i in outstanding), 2)

        # Uninvoiced time entries
        uninvoiced_entries = [e for e in all_entries if not e.get("invoiced")]
        uninvoiced_hours   = round(sum(e["hours"] for e in uninvoiced_entries), 2)
        uninvoiced_billed  = round(sum(
            e["hours"] * e.get("rate", project_map[e["project_id"]]["rate"])
            for e in uninvoiced_entries
            if e["project_id"] in project_map
        ), 2)

        # Uninvoiced expense records
        uninvoiced_expenses = [x for x in all_expenses if not x.get("invoiced")]
        uninvoiced_exp_records_total = round(sum(x["amount"] for x in uninvoiced_expenses), 2)

        # Outstanding expense invoices (sent/draft/late but not paid)
        outstanding_exp_invoices = [i for i in all_invoices
                                    if i.get("invoice_type") == "expense"
                                    and i.get("status") not in ("paid",)]
        outstanding_exp_invoices_total = round(sum(i.get("total", 0) for i in outstanding_exp_invoices), 2)

        uninvoiced_exp_total = round(uninvoiced_exp_records_total + outstanding_exp_invoices_total, 2)
        uninvoiced_exp_count = len(uninvoiced_expenses) + len(outstanding_exp_invoices)

        # Total pipeline = all three cards summed
        pipeline_total = round(outstanding_total + uninvoiced_billed + uninvoiced_exp_total, 2)

        # Outstanding invoices grouped by client for the table
        outstanding_by_client = {}
        for inv in outstanding:
            cid = inv.get("client_id", "")
            if cid not in outstanding_by_client:
                outstanding_by_client[cid] = {"client": client_map.get(cid, {"name": inv.get("client_name", "—")}), "invoices": [], "total": 0}
            outstanding_by_client[cid]["invoices"].append(inv)
            outstanding_by_client[cid]["total"] = round(outstanding_by_client[cid]["total"] + inv.get("total", 0), 2)

        # Outstanding expense invoices by client for expenses section
        exp_invoices_by_client = {}
        for inv in outstanding_exp_invoices:
            cid = inv.get("client_id", "")
            if cid not in exp_invoices_by_client:
                exp_invoices_by_client[cid] = {"client": client_map.get(cid, {"name": inv.get("client_name", "—")}), "invoices": [], "total": 0}
            exp_invoices_by_client[cid]["invoices"].append(inv)
            exp_invoices_by_client[cid]["total"] = round(exp_invoices_by_client[cid]["total"] + inv.get("total", 0), 2)

        # Uninvoiced time by client
        uninvoiced_by_client = {}
        for e in uninvoiced_entries:
            p = project_map.get(e["project_id"])
            if not p:
                continue
            cid = p.get("client_id", "")
            if cid not in uninvoiced_by_client:
                uninvoiced_by_client[cid] = {"client": client_map.get(cid, {"name": "—"}), "hours": 0, "amount": 0}
            uninvoiced_by_client[cid]["hours"]  = round(uninvoiced_by_client[cid]["hours"]  + e["hours"], 2)
            uninvoiced_by_client[cid]["amount"] = round(uninvoiced_by_client[cid]["amount"] + e["hours"] * e.get("rate", p["rate"]), 2)

        # Project health: active/on_hold projects with a budget set
        all_tasks_dash  = storage.load_tasks()
        all_pt_dash     = storage.load_project_tasks()
        task_map_dash   = {t["id"]: t for t in all_tasks_dash}
        project_health  = []
        for p in projects:
            status = p.get("status", "active" if p.get("active", True) else "complete")
            if status == "complete":
                continue
            bh = p.get("budget_hours")
            bd = p.get("budget_dollars")
            p_entries = [e for e in all_entries if e["project_id"] == p["id"]]
            used_h = round(sum(e["hours"] for e in p_entries), 2)
            used_d = round(sum(e["hours"] * e.get("rate", p["rate"]) for e in p_entries), 2)
            # Assigned tasks with budgets for this project
            p_pt = [r for r in all_pt_dash if r["project_id"] == p["id"]]
            if not bh and not bd:
                has_task_budgets = any(r.get("budget_hours") or r.get("budget_dollars") for r in p_pt)
                if not has_task_budgets:
                    continue
            entry = {
                "project": p,
                "client": client_map.get(p.get("client_id", ""), {"name": "—"}),
                "status": status,
                "used_hours": used_h,
                "used_dollars": used_d,
            }
            if bh:
                entry["hours_pct"]  = min(round(used_h / bh * 100, 1), 100)
                entry["hours_over"] = used_h > bh
            if bd:
                entry["dollars_pct"]  = min(round(used_d / bd * 100, 1), 100)
                entry["dollars_over"] = used_d > bd
            # Task-level health
            task_health = []
            for r in p_pt:
                tbh = r.get("budget_hours")
                tbd = r.get("budget_dollars")
                if not tbh and not tbd:
                    continue
                t = task_map_dash.get(r["task_id"])
                if not t:
                    continue
                t_entries = [e for e in p_entries if e.get("task_id") == t["id"]]
                t_used_h = round(sum(e["hours"] for e in t_entries), 2)
                t_used_d = round(sum(e["hours"] * e.get("rate", p["rate"]) for e in t_entries), 2)
                task_display = dict(t, budget_hours=tbh, budget_dollars=tbd)
                t_entry = {"task": task_display, "used_hours": t_used_h, "used_dollars": t_used_d}
                if tbh:
                    t_entry["hours_pct"]  = min(round(t_used_h / tbh * 100, 1), 100)
                    t_entry["hours_over"] = t_used_h > tbh
                if tbd:
                    t_entry["dollars_pct"]  = min(round(t_used_d / tbd * 100, 1), 100)
                    t_entry["dollars_over"] = t_used_d > tbd
                task_health.append(t_entry)
            entry["task_health"] = task_health
            project_health.append(entry)

        return render_template("dashboard.html",
            outstanding=outstanding,
            outstanding_total=outstanding_total,
            outstanding_by_client=list(outstanding_by_client.values()),
            uninvoiced_hours=uninvoiced_hours,
            uninvoiced_billed=uninvoiced_billed,
            uninvoiced_exp_total=uninvoiced_exp_total,
            uninvoiced_exp_count=uninvoiced_exp_count,
            exp_invoices_by_client=list(exp_invoices_by_client.values()),
            uninvoiced_by_client=list(uninvoiced_by_client.values()),
            pipeline_total=pipeline_total,
            project_health=project_health,
        )

    @app.route("/dashboard", methods=["GET", "POST"])
    def dashboard_old():
        if request.method == "POST":
            project_id = request.form["project_id"]
            date = request.form["date"]
            hours = float(request.form.get("hours", 0))
            description = request.form["description"].strip()
            if not project_id or hours <= 0:
                flash("Project and hours are required.", "error")
                return redirect(url_for("dashboard"))
            entries = storage.load_time_entries()
            entries.append({
                "id": storage.new_id(),
                "project_id": project_id,
                "date": date,
                "hours": hours,
                "description": description,
                "invoiced": False,
            })
            storage.save_time_entries(entries)
            flash("Time entry added.", "success")
            return redirect(url_for("dashboard"))

        today = dt_date.today().isoformat()
        projects = [p for p in storage.load_projects() if p["active"]]
        all_entries = storage.load_time_entries()
        today_entries = [e for e in all_entries if e["date"] == today]
        project_map = {p["id"]: p for p in storage.load_projects()}
        return render_template("dashboard.html", projects=projects,
                               today=today, today_entries=today_entries,
                               project_map=project_map)

    @app.route("/settings", methods=["GET", "POST"])
    def settings():
        if request.method == "POST":
            try:
                gst_rate = float(request.form.get("gst_rate", 0.05))
                smtp_port = int(request.form.get("smtp_port", 587))
                default_rate_str = request.form.get("default_rate", "").strip()
                default_rate = float(default_rate_str) if default_rate_str else None
            except ValueError:
                flash("GST rate, SMTP port, and default rate must be numbers.", "error")
                return redirect(url_for("settings"))
            s = storage.load_settings()
            s.update({
                "business_name": request.form["business_name"].strip(),
                "business_address": request.form.get("business_address", "").strip(),
                "gst_number": request.form["gst_number"].strip(),
                "gst_rate": gst_rate,
                "default_rate": default_rate,
                "smtp_host": request.form["smtp_host"].strip(),
                "smtp_port": smtp_port,
                "smtp_user": request.form["smtp_user"].strip(),
                "smtp_password": request.form["smtp_password"].strip(),
                "invoice_brand_color": request.form.get("invoice_brand_color", s.get("invoice_brand_color", "#1D1E1C")).strip() or "#1D1E1C",
                "invoice_bg_color": request.form.get("invoice_bg_color", s.get("invoice_bg_color", "#FFFFFF")).strip() or "#FFFFFF",
                "app_nav_color": request.form.get("app_nav_color", s.get("app_nav_color", "#1a1a2e")).strip() or "#1a1a2e",
                "app_accent_color": request.form.get("app_accent_color", s.get("app_accent_color", "#4f8ef7")).strip() or "#4f8ef7",
                "enable_tasks": request.form.get("enable_tasks") == "1",
                "enable_reports": request.form.get("enable_reports") == "1",
            })

            # Logo upload
            logo = request.files.get("invoice_logo")
            if logo and logo.filename:
                ext = logo.filename.rsplit(".", 1)[-1].lower()
                if ext in ("jpg", "jpeg", "png", "gif", "svg"):
                    logo_data = logo.read()
                    if len(logo_data) <= 5 * 1024 * 1024:
                        logo_dir = storage.DATA_DIR / "invoice_logos"
                        logo_dir.mkdir(parents=True, exist_ok=True)
                        fname = f"logo.{ext}"
                        (logo_dir / fname).write_bytes(logo_data)
                        s["invoice_logo_filename"] = fname
                    else:
                        flash("Logo must be under 5 MB.", "error")
                else:
                    flash("Logo must be JPG, PNG, GIF, or SVG.", "error")

            # Invoice logo removal
            if request.form.get("remove_logo"):
                s["invoice_logo_filename"] = None

            storage.save_settings(s)
            flash("Settings saved.", "success")
            return redirect(url_for("settings"))
        return render_template("settings.html", s=storage.load_settings())

    @app.route("/settings/logo")
    def settings_logo():
        s = storage.load_settings()
        fname = s.get("invoice_logo_filename")
        if not fname:
            return "", 404
        logo_path = storage.DATA_DIR / "invoice_logos" / fname
        if not logo_path.exists():
            return "", 404
        ext = fname.rsplit(".", 1)[-1].lower()
        mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "gif": "image/gif", "svg": "image/svg+xml"}.get(ext, "image/png")
        return send_file(str(logo_path), mimetype=mime)

    # ── Clients ──────────────────────────────────────────────────────────────

    @app.route("/clients")
    def clients():
        return render_template("clients.html", clients=storage.load_clients())

    @app.route("/clients/add", methods=["POST"])
    def clients_add():
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        address = request.form.get("address", "").strip()
        if not name:
            flash("Client name is required.", "error")
            return redirect(url_for("clients"))
        all_clients = storage.load_clients()
        all_clients.append({"id": storage.new_id(), "name": name, "email": email, "address": address})
        storage.save_clients(all_clients)
        flash(f"Client '{name}' added.", "success")
        return redirect(url_for("clients"))

    @app.route("/clients/<cid>/edit", methods=["POST"])
    def clients_edit(cid):
        all_clients = storage.load_clients()
        for c in all_clients:
            if c["id"] == cid:
                c["name"] = request.form.get("name", "").strip() or c["name"]
                c["email"] = request.form.get("email", "").strip()
                c["address"] = request.form.get("address", "").strip()
                storage.save_clients(all_clients)
                flash("Client updated.", "success")
                return redirect(url_for("clients"))
        flash("Client not found.", "error")
        return redirect(url_for("clients"))

    # ── Projects ─────────────────────────────────────────────────────────────

    @app.route("/projects")
    def projects():
        all_projects  = storage.load_projects()
        all_entries   = storage.load_time_entries()
        all_tasks     = storage.load_tasks()
        all_pt        = storage.load_project_tasks()
        project_map   = {p["id"]: p for p in all_projects}
        task_map      = {t["id"]: t for t in all_tasks}
        client_map    = {c["id"]: c for c in storage.load_clients()}
        project_stats = {}
        for e in all_entries:
            pid = e["project_id"]
            p = project_map.get(pid)
            if not p:
                continue
            if pid not in project_stats:
                project_stats[pid] = {"used_hours": 0, "used_dollars": 0}
            project_stats[pid]["used_hours"]   = round(project_stats[pid]["used_hours"]   + e["hours"], 2)
            project_stats[pid]["used_dollars"] = round(project_stats[pid]["used_dollars"] + e["hours"] * e.get("rate", p.get("rate", 0)), 2)
        # Build project_tasks: {pid: [merged task+assignment dicts]}
        project_tasks = {}
        for r in all_pt:
            t = task_map.get(r["task_id"])
            if not t:
                continue
            merged = {
                "id": t["id"],
                "pt_id": r["id"],
                "name": t["name"],
                "rate": r["rate_override"] if r.get("rate_override") is not None else t.get("rate"),
                "budget_hours": r.get("budget_hours"),
                "budget_dollars": r.get("budget_dollars"),
            }
            project_tasks.setdefault(r["project_id"], []).append(merged)
        # Compute which tasks are already assigned per project (for the assign dropdown)
        assigned_ids_per_project = {}
        for r in all_pt:
            assigned_ids_per_project.setdefault(r["project_id"], set()).add(r["task_id"])
        project_unassigned = {
            p["id"]: [t for t in all_tasks if t["id"] not in assigned_ids_per_project.get(p["id"], set())]
            for p in all_projects
        }

        # Organize projects hierarchically: top-level first, then their subs
        top_level = [p for p in all_projects if not p.get("parent_id")]
        subs_by_parent = {}
        for p in all_projects:
            if p.get("parent_id"):
                subs_by_parent.setdefault(p["parent_id"], []).append(p)
        ordered_projects = []
        top_ids = {p["id"] for p in top_level}
        for p in top_level:
            ordered_projects.append(p)
            for sp in subs_by_parent.get(p["id"], []):
                ordered_projects.append(sp)
        # Orphan subs (parent missing)
        ordered_ids = {p["id"] for p in ordered_projects}
        for p in all_projects:
            if p["id"] not in ordered_ids:
                ordered_projects.append(p)

        # Display status: sub-projects appear in their parent's tab
        project_display_status = {}
        for p in ordered_projects:
            natural = p.get("status", "active" if p.get("active", True) else "complete")
            if p.get("parent_id"):
                parent = project_map.get(p["parent_id"])
                project_display_status[p["id"]] = parent.get("status", "active" if parent.get("active", True) else "complete") if parent else natural
            else:
                project_display_status[p["id"]] = natural

        return render_template("projects.html",
            projects=ordered_projects,
            project_map=project_map,
            project_display_status=project_display_status,
            project_stats=project_stats,
            project_tasks=project_tasks,
            project_unassigned=project_unassigned,
            client_map=client_map,
            clients=storage.load_clients(),
            top_level_projects=top_level,
            all_tasks=all_tasks,
            settings=storage.load_settings()
        )

    @app.route("/projects/<pid>")
    def project_detail(pid):
        all_projects = storage.load_projects()
        project_map  = {p["id"]: p for p in all_projects}
        project = project_map.get(pid)
        if not project:
            flash("Project not found.", "error")
            return redirect(url_for("projects"))
        client_map = {c["id"]: c for c in storage.load_clients()}
        client = client_map.get(project.get("client_id", ""))
        task_map   = {t["id"]: t for t in storage.load_tasks()}
        all_entries = storage.load_time_entries()
        # Sub-projects of this project
        sub_projects = [p for p in all_projects if p.get("parent_id") == pid]
        # Collect relevant project ids: this project + its subs
        relevant_ids = {pid} | {sp["id"] for sp in sub_projects}
        entries = sorted(
            [e for e in all_entries if e.get("project_id") in relevant_ids],
            key=lambda e: e["date"], reverse=True)
        # Annotate entries with project and task name
        for e in entries:
            p = project_map.get(e["project_id"], {})
            e["_project_name"] = p.get("name", "")
            e["_task_name"]    = task_map.get(e.get("task_id", ""), {}).get("name", "")
            e["_amount"]       = round(e["hours"] * e.get("rate", p.get("rate", 0)), 2)
        total_hours  = round(sum(e["hours"] for e in entries), 2)
        total_amount = round(sum(e["_amount"] for e in entries), 2)
        invoiced_hours  = round(sum(e["hours"] for e in entries if e.get("invoiced")), 2)
        uninvoiced_hours = round(total_hours - invoiced_hours, 2)
        parent = project_map.get(project.get("parent_id")) if project.get("parent_id") else None
        return render_template("project_detail.html",
            project=project, parent=parent, client=client,
            sub_projects=sub_projects, entries=entries,
            total_hours=total_hours, total_amount=total_amount,
            invoiced_hours=invoiced_hours, uninvoiced_hours=uninvoiced_hours,
            project_map=project_map,
            settings=storage.load_settings())

    @app.route("/projects/add", methods=["POST"])
    def projects_add():
        name = request.form.get("name", "").strip()
        try:
            rate = float(request.form.get("rate", 0))
        except ValueError:
            flash("Rate must be a number.", "error")
            return redirect(url_for("projects"))
        if not name:
            flash("Project name is required.", "error")
            return redirect(url_for("projects"))
        client_id   = request.form.get("client_id", "").strip() or None
        parent_id   = request.form.get("parent_id", "").strip() or None
        status      = request.form.get("status", "active")
        due_date    = request.form.get("due_date", "").strip() or None
        try:
            bh_str = request.form.get("budget_hours", "").strip()
            budget_hours = float(bh_str) if bh_str else None
        except ValueError:
            budget_hours = None
        try:
            bd_str = request.form.get("budget_dollars", "").strip()
            budget_dollars = float(bd_str) if bd_str else None
        except ValueError:
            budget_dollars = None
        projects = storage.load_projects()
        if parent_id:
            parent = next((p for p in projects if p["id"] == parent_id), None)
            if parent:
                client_id = parent.get("client_id")
                rate = parent.get("rate", rate)
        projects.append({
            "id": storage.new_id(), "name": name, "rate": rate,
            "currency": "CAD", "active": status != "complete",
            "status": status, "client_id": client_id, "parent_id": parent_id,
            "due_date": due_date, "budget_hours": budget_hours, "budget_dollars": budget_dollars,
        })
        storage.save_projects(projects)
        flash(f"Project '{name}' added.", "success")
        return redirect(url_for("projects"))

    @app.route("/projects/<pid>/edit", methods=["POST"])
    def projects_edit(pid):
        projects = storage.load_projects()
        for p in projects:
            if p["id"] == pid:
                p["name"] = request.form.get("name", "").strip() or p["name"]
                try:
                    p["rate"] = float(request.form.get("rate", p["rate"]))
                except ValueError:
                    pass
                p["client_id"] = request.form.get("client_id", "").strip() or None
                p["parent_id"] = request.form.get("parent_id", "").strip() or None
                p["status"]    = request.form.get("status", p.get("status", "active"))
                p["active"]    = p["status"] != "complete"
                p["due_date"]  = request.form.get("due_date", "").strip() or None
                try:
                    bh_str = request.form.get("budget_hours", "").strip()
                    p["budget_hours"] = float(bh_str) if bh_str else None
                except ValueError:
                    pass
                try:
                    bd_str = request.form.get("budget_dollars", "").strip()
                    p["budget_dollars"] = float(bd_str) if bd_str else None
                except ValueError:
                    pass
                storage.save_projects(projects)
                flash("Project updated.", "success")
                return redirect(url_for("projects"))
        flash("Project not found.", "error")
        return redirect(url_for("projects"))

    @app.route("/projects/<pid>/deactivate", methods=["POST"])
    def projects_deactivate(pid):
        projects = storage.load_projects()
        for p in projects:
            if p["id"] == pid:
                p["active"] = False
                p["status"] = "complete"
                storage.save_projects(projects)
                flash("Project marked complete.", "success")
                return redirect(url_for("projects"))
        flash("Project not found.", "error")
        return redirect(url_for("projects"))

    @app.route("/projects/<pid>/reactivate", methods=["POST"])
    def projects_reactivate(pid):
        projects = storage.load_projects()
        for p in projects:
            if p["id"] == pid:
                p["active"] = True
                p["status"] = "active"
                storage.save_projects(projects)
                flash("Project reactivated.", "success")
                return redirect(url_for("projects"))
        flash("Project not found.", "error")
        return redirect(url_for("projects"))

    # ── Tasks (global library) ────────────────────────────────────────────────

    @app.route("/tasks")
    def tasks():
        all_tasks = storage.load_tasks()
        return render_template("tasks.html", tasks=all_tasks)

    @app.route("/tasks/add", methods=["POST"])
    def tasks_add():
        name = request.form.get("name", "").strip()
        if not name:
            flash("Task name is required.", "error")
            return redirect(url_for("tasks"))
        try:
            rate_str = request.form.get("rate", "").strip()
            rate = float(rate_str) if rate_str else None
        except ValueError:
            rate = None
        all_tasks = storage.load_tasks()
        all_tasks.append({"id": storage.new_id(), "name": name, "rate": rate})
        storage.save_tasks(all_tasks)
        flash(f"Task '{name}' added.", "success")
        return redirect(url_for("tasks"))

    @app.route("/tasks/<tid>/edit", methods=["POST"])
    def tasks_edit(tid):
        all_tasks = storage.load_tasks()
        for t in all_tasks:
            if t["id"] == tid:
                t["name"] = request.form.get("name", "").strip() or t["name"]
                try:
                    rate_str = request.form.get("rate", "").strip()
                    t["rate"] = float(rate_str) if rate_str else None
                except ValueError:
                    pass
                storage.save_tasks(all_tasks)
                flash("Task updated.", "success")
                return redirect(url_for("tasks"))
        flash("Task not found.", "error")
        return redirect(url_for("tasks"))

    @app.route("/tasks/<tid>/delete", methods=["POST"])
    def tasks_delete(tid):
        all_tasks = [t for t in storage.load_tasks() if t["id"] != tid]
        storage.save_tasks(all_tasks)
        # Remove all project assignments for this task
        pt = [r for r in storage.load_project_tasks() if r["task_id"] != tid]
        storage.save_project_tasks(pt)
        flash("Task deleted.", "success")
        return redirect(url_for("tasks"))

    # ── Project-Task assignments ──────────────────────────────────────────────

    @app.route("/projects/<pid>/tasks/assign", methods=["POST"])
    def project_tasks_assign(pid):
        task_id = request.form.get("task_id", "").strip()
        if not task_id:
            flash("Select a task to assign.", "error")
            return redirect(url_for("projects"))
        pt = storage.load_project_tasks()
        if any(r["project_id"] == pid and r["task_id"] == task_id for r in pt):
            flash("Task already assigned to this project.", "error")
            return redirect(url_for("projects"))
        try:
            rate_str = request.form.get("rate_override", "").strip()
            rate_override = float(rate_str) if rate_str else None
        except ValueError:
            rate_override = None
        try:
            bh_str = request.form.get("budget_hours", "").strip()
            budget_hours = float(bh_str) if bh_str else None
        except ValueError:
            budget_hours = None
        try:
            bd_str = request.form.get("budget_dollars", "").strip()
            budget_dollars = float(bd_str) if bd_str else None
        except ValueError:
            budget_dollars = None
        pt.append({
            "id": storage.new_id(),
            "project_id": pid,
            "task_id": task_id,
            "rate_override": rate_override,
            "budget_hours": budget_hours,
            "budget_dollars": budget_dollars,
        })
        storage.save_project_tasks(pt)
        flash("Task assigned.", "success")
        return redirect(url_for("projects"))

    @app.route("/project-tasks/<ptid>/edit", methods=["POST"])
    def project_tasks_edit(ptid):
        pt = storage.load_project_tasks()
        for r in pt:
            if r["id"] == ptid:
                try:
                    rate_str = request.form.get("rate_override", "").strip()
                    r["rate_override"] = float(rate_str) if rate_str else None
                except ValueError:
                    pass
                try:
                    bh_str = request.form.get("budget_hours", "").strip()
                    r["budget_hours"] = float(bh_str) if bh_str else None
                except ValueError:
                    pass
                try:
                    bd_str = request.form.get("budget_dollars", "").strip()
                    r["budget_dollars"] = float(bd_str) if bd_str else None
                except ValueError:
                    pass
                storage.save_project_tasks(pt)
                flash("Task assignment updated.", "success")
                return redirect(url_for("projects"))
        flash("Assignment not found.", "error")
        return redirect(url_for("projects"))

    @app.route("/project-tasks/<ptid>/remove", methods=["POST"])
    def project_tasks_remove(ptid):
        pt = [r for r in storage.load_project_tasks() if r["id"] != ptid]
        storage.save_project_tasks(pt)
        flash("Task removed from project.", "success")
        return redirect(url_for("projects"))

    @app.route("/time")
    def time_entries():
        import calendar as cal_mod
        today = dt_date.today()
        view = request.args.get("view", "day")

        all_projects = storage.load_projects()
        active_projects = [p for p in all_projects if p["active"]]
        project_map = {p["id"]: p for p in all_projects}
        all_entries = storage.load_time_entries()
        all_tasks = storage.load_tasks()
        all_pt    = storage.load_project_tasks()
        task_map  = {t["id"]: t for t in all_tasks}
        # group assigned tasks by project_id for JS (use rate_override if set)
        project_tasks = {}
        for r in all_pt:
            t = task_map.get(r["task_id"])
            if not t:
                continue
            rate = r["rate_override"] if r.get("rate_override") is not None else t.get("rate")
            project_tasks.setdefault(r["project_id"], []).append({"id": t["id"], "name": t["name"], "rate": rate})

        # Build project groups for optgroup rendering in dropdowns
        active_top = [p for p in active_projects if not p.get("parent_id")]
        active_subs = [p for p in active_projects if p.get("parent_id")]
        subs_by_parent = {}
        for sp in active_subs:
            subs_by_parent.setdefault(sp["parent_id"], []).append(sp)
        active_top_ids = {p["id"] for p in active_top}
        project_groups = []
        for p in active_top:
            project_groups.append({"project": p, "subs": subs_by_parent.get(p["id"], [])})
        for sp in active_subs:
            if sp["parent_id"] not in active_top_ids:
                project_groups.append({"project": sp, "subs": []})

        # ── shared grid_json for popup modal (all entries keyed by project+date) ──
        def make_grid_json(entries, dates):
            g = {p["id"]: {d.isoformat(): [] for d in dates} for p in all_projects}
            for e in entries:
                pid = e["project_id"]
                if pid in g and e["date"] in g[pid]:
                    g[pid][e["date"]].append(e)
            return {
                pid: {
                    ds: [{"id": e["id"], "hours": e["hours"],
                          "description": e.get("description", ""),
                          "rate": e.get("rate"),
                          "task_id": e.get("task_id"),
                          "invoiced": e.get("invoiced", False)}
                         for e in el]
                    for ds, el in dmap.items()
                }
                for pid, dmap in g.items()
            }

        # ── WEEK ──────────────────────────────────────────────────────────────
        week_str = request.args.get("week", "")
        if week_str:
            try:
                week_start = dt_date.fromisoformat(week_str)
                week_start = week_start - timedelta(days=week_start.weekday())
            except ValueError:
                week_start = today - timedelta(days=today.weekday())
        else:
            week_start = today - timedelta(days=today.weekday())

        week_dates = [week_start + timedelta(days=i) for i in range(7)]
        week_end = week_dates[-1]
        prev_week = (week_start - timedelta(days=7)).isoformat()
        next_week = (week_start + timedelta(days=7)).isoformat()
        today_week = (today - timedelta(days=today.weekday())).isoformat()
        d0, d6 = week_dates[0], week_dates[-1]
        week_label = (f"{d0.strftime('%b')} {d0.day} – {d6.day}, {d6.year}"
                      if d0.month == d6.month
                      else f"{d0.strftime('%b')} {d0.day} – {d6.strftime('%b')} {d6.day}, {d6.year}")
        week_entries = [e for e in all_entries
                        if week_start.isoformat() <= e["date"] <= week_end.isoformat()]
        grid = {p["id"]: {d.isoformat(): [] for d in week_dates} for p in all_projects}
        for e in week_entries:
            pid = e["project_id"]
            if pid in grid and e["date"] in grid[pid]:
                grid[pid][e["date"]].append(e)
        row_totals = {p["id"]: round(
            sum(e["hours"] for d in week_dates for e in grid[p["id"]][d.isoformat()]), 2)
            for p in all_projects}
        daily_totals_week = {d.isoformat(): round(
            sum(e["hours"] for e in week_entries if e["date"] == d.isoformat()), 2)
            for d in week_dates}
        week_total = round(sum(daily_totals_week.values()), 2)
        grid_json = make_grid_json(week_entries, week_dates)

        # ── DAY ───────────────────────────────────────────────────────────────
        day_str = request.args.get("date", today.isoformat())
        try:
            day_date = dt_date.fromisoformat(day_str)
        except ValueError:
            day_date = today
        prev_day = (day_date - timedelta(days=1)).isoformat()
        next_day = (day_date + timedelta(days=1)).isoformat()
        day_entries_raw = [e for e in all_entries if e["date"] == day_date.isoformat()]
        day_grid = {p["id"]: [] for p in all_projects}
        for e in day_entries_raw:
            if e["project_id"] in day_grid:
                day_grid[e["project_id"]].append(e)
        day_total = round(sum(e["hours"] for e in day_entries_raw), 2)
        day_grid_json = make_grid_json(day_entries_raw, [day_date])

        # ── MONTH ─────────────────────────────────────────────────────────────
        month_str = request.args.get("month", today.strftime("%Y-%m"))
        try:
            m_year, m_month = int(month_str[:4]), int(month_str[5:7])
        except (ValueError, IndexError):
            m_year, m_month = today.year, today.month
        first_day = dt_date(m_year, m_month, 1)
        last_day = dt_date(m_year, m_month, cal_mod.monthrange(m_year, m_month)[1])
        prev_month_d = first_day - timedelta(days=1)
        next_month_d = last_day + timedelta(days=1)
        prev_month = prev_month_d.strftime("%Y-%m")
        next_month = next_month_d.strftime("%Y-%m")
        month_label = first_day.strftime("%B %Y")
        # Calendar weeks (Mon-based, pad with None)
        cal_start = first_day - timedelta(days=first_day.weekday())
        cal_end   = last_day + timedelta(days=(6 - last_day.weekday()))
        cal_days  = [cal_start + timedelta(days=i) for i in range((cal_end - cal_start).days + 1)]
        cal_weeks = [cal_days[i:i+7] for i in range(0, len(cal_days), 7)]
        month_entries = [e for e in all_entries
                         if first_day.isoformat() <= e["date"] <= last_day.isoformat()]
        daily_totals_month = {}
        for e in month_entries:
            daily_totals_month[e["date"]] = round(
                daily_totals_month.get(e["date"], 0) + e["hours"], 2)
        month_total = round(sum(daily_totals_month.values()), 2)

        # ── LIST ──────────────────────────────────────────────────────────────
        lf_project  = request.args.get("lf_project", "")
        lf_from     = request.args.get("lf_from", "")
        lf_to       = request.args.get("lf_to", "")
        lf_invoiced = request.args.get("lf_invoiced", "")
        list_entries = list(all_entries)
        if lf_project:
            list_entries = [e for e in list_entries if e["project_id"] == lf_project]
        if lf_from:
            list_entries = [e for e in list_entries if e["date"] >= lf_from]
        if lf_to:
            list_entries = [e for e in list_entries if e["date"] <= lf_to]
        if lf_invoiced == "0":
            list_entries = [e for e in list_entries if not e.get("invoiced")]
        elif lf_invoiced == "1":
            list_entries = [e for e in list_entries if e.get("invoiced")]
        list_entries = sorted(list_entries, key=lambda e: e["date"], reverse=True)
        list_total_hours  = round(sum(e["hours"] for e in list_entries), 2)
        list_total_amount = round(sum(
            e["hours"] * e.get("rate", project_map.get(e.get("project_id",""), {}).get("rate", 0))
            for e in list_entries), 2)

        return render_template("time.html",
                               view=view,
                               projects=all_projects,
                               project_groups=project_groups,
                               project_map=project_map,
                               project_tasks=project_tasks,
                               task_map=task_map,
                               # week
                               week_dates=week_dates,
                               week_start=week_start,
                               week_label=week_label,
                               prev_week=prev_week,
                               next_week=next_week,
                               today=today.isoformat(),
                               today_week=today_week,
                               grid=grid,
                               grid_json=grid_json,
                               row_totals=row_totals,
                               daily_totals=daily_totals_week,
                               week_total=week_total,
                               # day
                               day_date=day_date,
                               prev_day=prev_day,
                               next_day=next_day,
                               day_grid=day_grid,
                               day_grid_json=day_grid_json,
                               day_total=day_total,
                               # month
                               cal_weeks=cal_weeks,
                               month_label=month_label,
                               prev_month=prev_month,
                               next_month=next_month,
                               daily_totals_month=daily_totals_month,
                               month_total=month_total,
                               m_year=m_year,
                               m_month=m_month,
                               # list
                               list_entries=list_entries,
                               list_total_hours=list_total_hours,
                               list_total_amount=list_total_amount,
                               lf_project=lf_project,
                               lf_from=lf_from,
                               lf_to=lf_to,
                               lf_invoiced=lf_invoiced)

    def _week_redirect(date_str):
        """Redirect to the weekly view containing date_str."""
        try:
            d = dt_date.fromisoformat(date_str)
            week_start = d - timedelta(days=d.weekday())
            return redirect(url_for("time_entries", week=week_start.isoformat()))
        except (ValueError, TypeError):
            return redirect(url_for("time_entries"))

    @app.route("/time/add", methods=["POST"])
    def time_add():
        project_id = request.form.get("project_id", "").strip()
        date = request.form.get("date", "").strip()
        try:
            hours = float(request.form.get("hours", 0))
        except ValueError:
            flash("Hours must be a number.", "error")
            return redirect(url_for("time_entries"))
        if not project_id or hours <= 0:
            flash("Project and hours required.", "error")
            return redirect(url_for("time_entries"))
        if not date:
            flash("Date is required.", "error")
            return redirect(url_for("time_entries"))
        all_projects = storage.load_projects()
        project_map = {p["id"]: p for p in all_projects}
        if any(p.get("parent_id") == project_id for p in all_projects):
            flash("Cannot log time to a parent project — select a sub-project instead.", "error")
            return redirect(url_for("time_entries"))
        default_rate = project_map.get(project_id, {}).get("rate", 0)
        try:
            rate = float(request.form.get("rate", default_rate))
        except ValueError:
            rate = default_rate
        entries = storage.load_time_entries()
        task_id = request.form.get("task_id", "").strip() or None
        entries.append({
            "id": storage.new_id(),
            "project_id": project_id,
            "task_id": task_id,
            "date": date,
            "hours": hours,
            "rate": rate,
            "description": request.form.get("description", "").strip(),
            "invoiced": False,
        })
        storage.save_time_entries(entries)
        return _week_redirect(date)

    @app.route("/time/<eid>/edit", methods=["GET", "POST"])
    def time_edit(eid):
        entries = storage.load_time_entries()
        entry = next((e for e in entries if e["id"] == eid), None)
        if not entry:
            flash("Entry not found.", "error")
            return redirect(url_for("time_entries"))
        if entry["invoiced"]:
            flash("Cannot edit an invoiced entry.", "error")
            return _week_redirect(entry.get("date", ""))
        if request.method == "POST":
            new_project_id = request.form.get("project_id", entry["project_id"])
            all_projects_edit = storage.load_projects()
            if any(p.get("parent_id") == new_project_id for p in all_projects_edit):
                flash("Cannot log time to a parent project — select a sub-project instead.", "error")
                return _week_redirect(entry.get("date", ""))
            entry["project_id"] = new_project_id
            entry["date"] = request.form.get("date", entry["date"]).strip() or entry["date"]
            try:
                hours = float(request.form.get("hours", entry["hours"]))
            except ValueError:
                flash("Hours must be a number.", "error")
                return _week_redirect(entry.get("date", ""))
            if hours <= 0:
                flash("Hours must be greater than zero.", "error")
                return _week_redirect(entry.get("date", ""))
            entry["hours"] = hours
            entry["description"] = request.form.get("description", "").strip()
            entry["task_id"] = request.form.get("task_id", "").strip() or None
            project_map = {p["id"]: p for p in all_projects_edit}
            default_rate = project_map.get(entry["project_id"], {}).get("rate", 0)
            try:
                entry["rate"] = float(request.form.get("rate", default_rate))
            except ValueError:
                pass
            storage.save_time_entries(entries)
            return _week_redirect(entry["date"])
        return _week_redirect(entry.get("date", ""))

    @app.route("/time/<eid>/delete", methods=["POST"])
    def time_delete(eid):
        entries = storage.load_time_entries()
        entry = next((e for e in entries if e["id"] == eid), None)
        if entry and entry["invoiced"]:
            flash("Cannot delete an invoiced entry.", "error")
            return _week_redirect(entry.get("date", ""))
        date = entry["date"] if entry else ""
        entries = [e for e in entries if e["id"] != eid]
        storage.save_time_entries(entries)
        return _week_redirect(date)

    # ── Expenses (by client) ──────────────────────────────────────────────────

    @app.route("/expenses")
    def expenses():
        _migrate_expenses_to_client()
        items = storage.load_expenses()
        clients = storage.load_clients()
        client_map = {c["id"]: c for c in clients}
        cf = request.args.get("client_id", "")
        df = request.args.get("date_from", "")
        dt = request.args.get("date_to", "")
        if cf:
            items = [x for x in items if x.get("client_id") == cf]
        if df:
            items = [x for x in items if x["date"] >= df]
        if dt:
            items = [x for x in items if x["date"] <= dt]
        items = sorted(items, key=lambda x: x["date"], reverse=True)
        return render_template("expenses.html", expenses=items, clients=clients,
                               client_map=client_map, cf=cf, df=df, dt=dt)

    @app.route("/expenses/add", methods=["POST"])
    def expenses_add():
        _migrate_expenses_to_client()
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        client_id = request.form.get("client_id", "").strip()
        date = request.form.get("date", "").strip()
        try:
            amount = float(request.form.get("amount", 0))
        except ValueError:
            flash("Amount must be a number.", "error")
            return redirect(url_for("expenses"))
        if not client_id or amount <= 0:
            flash("Client and amount required.", "error")
            return redirect(url_for("expenses"))
        if not date:
            flash("Date is required.", "error")
            return redirect(url_for("expenses"))

        pdf_filename = None
        receipt = request.files.get("receipt")
        if receipt and receipt.filename and receipt.filename.lower().endswith(".pdf") \
                and receipt.mimetype == "application/pdf":
            expense_pdf_dir.mkdir(parents=True, exist_ok=True)
            pdf_filename = storage.new_id() + ".pdf"
            receipt.save(str(expense_pdf_dir / pdf_filename))

        items = storage.load_expenses()
        items.append({
            "id": storage.new_id(),
            "client_id": client_id,
            "date": date,
            "amount": amount,
            "description": request.form.get("description", "").strip(),
            "gst": request.form.get("gst") == "1",
            "pdf_filename": pdf_filename,
            "invoiced": False,
        })
        storage.save_expenses(items)
        flash("Expense added.", "success")
        return redirect(url_for("expenses"))

    @app.route("/expenses/<xid>/edit", methods=["GET", "POST"])
    def expenses_edit(xid):
        _migrate_expenses_to_client()
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        items = storage.load_expenses()
        item = next((x for x in items if x["id"] == xid), None)
        if not item:
            flash("Expense not found.", "error")
            return redirect(url_for("expenses"))
        if item["invoiced"]:
            flash("Cannot edit an invoiced expense.", "error")
            return redirect(url_for("expenses"))
        if request.method == "POST":
            item["client_id"] = request.form.get("client_id", "").strip()
            item["date"] = request.form.get("date", item["date"]).strip() or item["date"]
            try:
                amount = float(request.form.get("amount", item["amount"]))
            except ValueError:
                flash("Amount must be a number.", "error")
                return redirect(url_for("expenses"))
            if amount <= 0:
                flash("Amount must be greater than zero.", "error")
                return redirect(url_for("expenses"))
            item["amount"] = amount
            item["description"] = request.form.get("description", "").strip()
            item["gst"] = request.form.get("gst") == "1"
            receipt = request.files.get("receipt")
            if receipt and receipt.filename and receipt.filename.lower().endswith(".pdf") \
                    and receipt.mimetype == "application/pdf":
                if item.get("pdf_filename"):
                    old_path = expense_pdf_dir / item["pdf_filename"]
                    if old_path.exists():
                        old_path.unlink()
                expense_pdf_dir.mkdir(parents=True, exist_ok=True)
                pdf_filename = storage.new_id() + ".pdf"
                receipt.save(str(expense_pdf_dir / pdf_filename))
                item["pdf_filename"] = pdf_filename
            storage.save_expenses(items)
            flash("Expense updated.", "success")
            return redirect(url_for("expenses"))
        clients = storage.load_clients()
        client_map = {c["id"]: c for c in clients}
        return render_template("expenses.html", edit_expense=item,
                               expenses=sorted(items, key=lambda x: x["date"], reverse=True),
                               clients=clients,
                               client_map=client_map,
                               cf="", df="", dt="")

    @app.route("/expenses/<xid>/delete", methods=["POST"])
    def expenses_delete(xid):
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        items = storage.load_expenses()
        item = next((x for x in items if x["id"] == xid), None)
        if item and item["invoiced"]:
            flash("Cannot delete an invoiced expense.", "error")
            return redirect(url_for("expenses"))
        if item and item.get("pdf_filename"):
            pdf_path = expense_pdf_dir / item["pdf_filename"]
            if pdf_path.exists():
                pdf_path.unlink()
        items = [x for x in items if x["id"] != xid]
        storage.save_expenses(items)
        flash("Expense deleted.", "success")
        return redirect(url_for("expenses"))

    @app.route("/expenses/<xid>/receipt")
    def expenses_receipt(xid):
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        items = storage.load_expenses()
        item = next((x for x in items if x["id"] == xid), None)
        if not item or not item.get("pdf_filename"):
            flash("No receipt found.", "error")
            return redirect(url_for("expenses"))
        pdf_path = expense_pdf_dir / item["pdf_filename"]
        if not pdf_path.exists():
            flash("Receipt file missing.", "error")
            return redirect(url_for("expenses"))
        return send_file(str(pdf_path), download_name=f"receipt-{item['date']}.pdf",
                         as_attachment=True, mimetype="application/pdf")

    # ── Invoices ──────────────────────────────────────────────────────────────

    def _next_invoice_number(prefix: str) -> str:
        """Return the next unique invoice number for a given prefix (INV or EXP),
        storing a monotonically-increasing counter in settings so numbers are
        never reused even after deletions."""
        s = storage.load_settings()
        key = f"last_{prefix.lower()}_num"
        # Seed from existing invoices on first use so we don't go backwards
        existing = storage.load_invoices()
        existing_max = max(
            (int(i["invoice_number"].split("-")[1])
             for i in existing
             if i.get("invoice_number", "").startswith(f"{prefix}-")
             and i["invoice_number"].split("-")[1].isdigit()),
            default=399
        )
        counter = max(int(s.get(key, 399)), existing_max)
        counter += 1
        s[key] = counter
        storage.save_settings(s)
        return f"{prefix}-{counter:03d}"

    @app.route("/invoices")
    def invoices():
        _migrate_expenses_to_client()
        client_id = request.args.get("client_id", "")
        all_clients = storage.load_clients()
        client_map = {c["id"]: c for c in all_clients}
        projects = storage.load_projects()
        project_map = {p["id"]: p for p in projects}
        all_entries = storage.load_time_entries()
        all_expenses = storage.load_expenses()
        all_invoices = sorted(storage.load_invoices(), key=lambda i: i.get("issued_date", ""), reverse=True)
        # Migrate invoices that predate the status field
        for inv in all_invoices:
            if "status" not in inv:
                inv["status"] = "sent" if inv.get("sent") else "draft"
            # Backfill entries_date for invoices generated before this field existed
            if "entries_date" not in inv:
                dates = []
                for li in inv.get("line_items", []):
                    if li.get("date"):
                        dates.append(li["date"])
                    elif li.get("description") and " to " in li["description"]:
                        dates.append(li["description"].split(" to ")[-1].strip())
                    elif li.get("description") and li["description"][:4].isdigit():
                        dates.append(li["description"].strip())
                inv["entries_date"] = max(dates) if dates else inv.get("issued_date", "")

        time_preview = None
        expense_preview = None
        if client_id:
            client = client_map.get(client_id)
            client_projects = [p for p in projects if p.get("client_id") == client_id]
            settings_data = storage.load_settings()
            gst_rate = settings_data.get("gst_rate", 0.05)

            # Time invoice preview (time entries only)
            time_sections = []
            for p in client_projects:
                time_entries = sorted(
                    [e for e in all_entries if e["project_id"] == p["id"] and not e["invoiced"]],
                    key=lambda e: e["date"])
                if time_entries:
                    parent_p = project_map.get(p.get("parent_id")) if p.get("parent_id") else None
                    display_name = f"{parent_p['name']} › {p['name']}" if parent_p else p["name"]
                    time_sections.append({"project": p, "display_name": display_name, "time_entries": time_entries})
            if time_sections:
                subtotal = round(
                    sum(sum(e["hours"] * e.get("rate", s["project"]["rate"]) for e in s["time_entries"])
                        for s in time_sections), 2)
                gst = round(subtotal * gst_rate, 2)
                time_preview = {
                    "client": client,
                    "sections": time_sections,
                    "subtotal": subtotal,
                    "gst": gst,
                    "total": round(subtotal + gst, 2),
                }

            # Expense invoice preview (expenses only)
            client_expenses = sorted(
                [x for x in all_expenses if x.get("client_id") == client_id and not x["invoiced"]],
                key=lambda x: x["date"])
            if client_expenses:
                subtotal = round(sum(x["amount"] for x in client_expenses), 2)
                gst = round(sum(x["amount"] * gst_rate for x in client_expenses if x.get("gst", True)), 2)
                expense_preview = {
                    "client": client,
                    "expenses": client_expenses,
                    "receipt_count": sum(1 for x in client_expenses if x.get("pdf_filename")),
                    "subtotal": subtotal,
                    "gst": gst,
                    "total": round(subtotal + gst, 2),
                }

        return render_template("invoices.html", clients=all_clients, client_map=client_map,
                               project_map=project_map, invoices=all_invoices,
                               time_preview=time_preview, expense_preview=expense_preview,
                               selected_client=client_id)

    @app.route("/invoices/generate", methods=["POST"])
    def invoices_generate():
        _migrate_expenses_to_client()
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        client_id = request.form["client_id"]
        detail_level = request.form.get("detail_level", "detailed")
        settings_data = storage.load_settings()
        all_clients = storage.load_clients()
        client_map = {c["id"]: c for c in all_clients}
        client = client_map.get(client_id)
        if not client:
            flash("Client not found.", "error")
            return redirect(url_for("invoices"))

        projects = storage.load_projects()
        project_map_inv = {p["id"]: p for p in projects}
        client_projects = [p for p in projects if p.get("client_id") == client_id]

        all_entries = storage.load_time_entries()
        all_expenses = storage.load_expenses()
        task_map_inv = {t["id"]: t for t in storage.load_tasks()}

        line_items = []
        entry_ids_invoiced = set()

        for p in client_projects:
            parent_p = project_map_inv.get(p.get("parent_id")) if p.get("parent_id") else None
            project_display_name = f"{parent_p['name']} › {p['name']}" if parent_p else p["name"]
            uninvoiced_time = sorted(
                [e for e in all_entries if e["project_id"] == p["id"] and not e["invoiced"]],
                key=lambda e: e["date"])

            if not uninvoiced_time:
                continue

            if detail_level == "summary":
                total_hours = round(sum(e["hours"] for e in uninvoiced_time), 2)
                effective_rate = uninvoiced_time[0].get("rate", p["rate"]) if len(set(e.get("rate", p["rate"]) for e in uninvoiced_time)) == 1 else p["rate"]
                amount = round(sum(e["hours"] * e.get("rate", p["rate"]) for e in uninvoiced_time), 2)
                entry_ids_str = ",".join(e["id"] for e in uninvoiced_time)
                dates = sorted(e["date"] for e in uninvoiced_time)
                summary_description = dates[0] if dates[0] == dates[-1] else f"{dates[0]} to {dates[-1]}"
                line_items.append({
                    "type": "time",
                    "description": summary_description,
                    "project_name": project_display_name,
                    "hours": total_hours,
                    "rate": effective_rate,
                    "amount": amount,
                    "entry_id": "",
                    "entry_ids": entry_ids_str,
                })
                for e in uninvoiced_time:
                    entry_ids_invoiced.add(e["id"])
            else:
                for e in uninvoiced_time:
                    entry_rate = e.get("rate", p["rate"])
                    amount = round(e["hours"] * entry_rate, 2)
                    task_name = task_map_inv.get(e.get("task_id", ""), {}).get("name", "")
                    line_items.append({
                        "type": "time",
                        "date": e["date"],
                        "description": e["description"],
                        "task_name": task_name,
                        "project_name": project_display_name,
                        "hours": e["hours"],
                        "rate": entry_rate,
                        "amount": amount,
                        "entry_id": e["id"],
                    })
                    entry_ids_invoiced.add(e["id"])

        if not line_items:
            flash("No uninvoiced time entries for this client.", "error")
            return redirect(url_for("invoices"))

        all_invoiced_dates = [e["date"] for e in all_entries if e["id"] in entry_ids_invoiced]
        entries_date = max(all_invoiced_dates) if all_invoiced_dates else dt_date.today().isoformat()

        subtotal = round(sum(li["amount"] for li in line_items), 2)
        gst_rate = settings_data.get("gst_rate", 0.05)
        gst = round(subtotal * gst_rate, 2)
        total = round(subtotal + gst, 2)

        invoice_number = _next_invoice_number("INV")
        existing = storage.load_invoices()

        invoice = {
            "id": storage.new_id(),
            "invoice_number": invoice_number,
            "invoice_type": "time",
            "client_id": client["id"],
            "client_name": client["name"],
            "client_email": client.get("email", ""),
            "client_address": client.get("address", ""),
            "issued_date": dt_date.today().isoformat(),
            "entries_date": entries_date,
            "subtotal": subtotal,
            "gst": gst,
            "total": total,
            "status": "draft",
            "sent": False,
            "detail_level": detail_level,
            "line_items": line_items,
        }

        existing.append(invoice)
        storage.save_invoices(existing)

        for e in all_entries:
            if e["id"] in entry_ids_invoiced:
                e["invoiced"] = True
        storage.save_time_entries(all_entries)

        flash(f"Invoice {invoice_number} generated.", "success")
        return redirect(url_for("invoices"))

    @app.route("/invoices/generate-expense", methods=["POST"])
    def invoices_generate_expense():
        _migrate_expenses_to_client()
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        client_id = request.form["client_id"]
        settings_data = storage.load_settings()
        all_clients = storage.load_clients()
        client_map = {c["id"]: c for c in all_clients}
        client = client_map.get(client_id)
        if not client:
            flash("Client not found.", "error")
            return redirect(url_for("invoices"))

        all_expenses = storage.load_expenses()
        uninvoiced_exp = sorted(
            [x for x in all_expenses if x.get("client_id") == client_id and not x["invoiced"]],
            key=lambda x: x["date"])

        if not uninvoiced_exp:
            flash("No uninvoiced expenses for this client.", "error")
            return redirect(url_for("invoices"))

        gst_rate = settings_data.get("gst_rate", 0.05)
        line_items = []
        expense_ids_invoiced = set()
        for x in uninvoiced_exp:
            include_gst = x.get("gst", True)
            item_gst = round(x["amount"] * gst_rate, 2) if include_gst else 0.0
            line_items.append({
                "type": "expense",
                "date": x["date"],
                "description": x["description"],
                "project_name": "",
                "amount": x["amount"],
                "item_gst": item_gst,
                "pdf_filename": x.get("pdf_filename"),
                "expense_id": x["id"],
            })
            expense_ids_invoiced.add(x["id"])

        subtotal = round(sum(li["amount"] for li in line_items), 2)
        gst = round(sum(li["item_gst"] for li in line_items), 2)
        total = round(subtotal + gst, 2)

        invoice_number = _next_invoice_number("EXP")
        existing = storage.load_invoices()

        invoice = {
            "id": storage.new_id(),
            "invoice_number": invoice_number,
            "invoice_type": "expense",
            "client_id": client["id"],
            "client_name": client["name"],
            "client_email": client.get("email", ""),
            "client_address": client.get("address", ""),
            "issued_date": dt_date.today().isoformat(),
            "subtotal": subtotal,
            "gst": gst,
            "total": total,
            "status": "draft",
            "sent": False,
            "line_items": line_items,
        }

        existing.append(invoice)
        storage.save_invoices(existing)

        for x in all_expenses:
            if x["id"] in expense_ids_invoiced:
                x["invoiced"] = True
        storage.save_expenses(all_expenses)

        flash(f"Expense invoice {invoice_number} generated.", "success")
        return redirect(url_for("invoices"))

    @app.route("/invoices/create-blank", methods=["POST"])
    def invoices_create_blank():
        client_id = request.form.get("client_id", "").strip()
        settings_data = storage.load_settings()
        all_clients = storage.load_clients()
        client_map = {c["id"]: c for c in all_clients}
        client = client_map.get(client_id)
        if not client:
            flash("Please select a client.", "error")
            return redirect(url_for("invoices"))

        invoice_number = _next_invoice_number("INV")
        existing = storage.load_invoices()
        gst_rate = settings_data.get("gst_rate", 0.05)
        invoice = {
            "id": storage.new_id(),
            "invoice_number": invoice_number,
            "invoice_type": "time",
            "client_id": client["id"],
            "client_name": client["name"],
            "client_email": client.get("email", ""),
            "client_address": client.get("address", ""),
            "issued_date": dt_date.today().isoformat(),
            "subtotal": 0.0,
            "gst": 0.0,
            "total": 0.0,
            "status": "draft",
            "sent": False,
            "line_items": [],
        }
        existing.append(invoice)
        storage.save_invoices(existing)
        return redirect(url_for("invoices_edit", inv_id=invoice["id"]))

    @app.route("/invoices/<inv_id>/edit", methods=["GET", "POST"])
    def invoices_edit(inv_id):
        inv_list = storage.load_invoices()
        invoice = next((i for i in inv_list if i["id"] == inv_id), None)
        if not invoice:
            flash("Invoice not found.", "error")
            return redirect(url_for("invoices"))

        if request.method == "POST":
            all_entries = storage.load_time_entries()
            all_expenses = storage.load_expenses()
            settings_data = storage.load_settings()

            invoice["invoice_number"] = request.form.get("invoice_number", invoice["invoice_number"]).strip() or invoice["invoice_number"]
            invoice["issued_date"] = request.form.get("issued_date", invoice["issued_date"]).strip() or invoice["issued_date"]
            invoice["subject"] = request.form.get("subject", "").strip()

            # Process existing line items
            new_line_items = []
            i = 0
            while f"item_{i}_amount" in request.form:
                if request.form.get(f"item_{i}_remove"):
                    # Un-invoice the entry/expense so it can be rebilled
                    entry_id = request.form.get(f"item_{i}_entry_id", "")
                    entry_ids_str = request.form.get(f"item_{i}_entry_ids", "")
                    expense_id = request.form.get(f"item_{i}_expense_id", "")
                    ids_to_uninvoice = set()
                    if entry_id:
                        ids_to_uninvoice.add(entry_id)
                    if entry_ids_str:
                        ids_to_uninvoice.update(entry_ids_str.split(","))
                    for e in all_entries:
                        if e["id"] in ids_to_uninvoice:
                            e["invoiced"] = False
                    for x in all_expenses:
                        if x["id"] == expense_id:
                            x["invoiced"] = False
                else:
                    item = {k: request.form.get(f"item_{i}_{k}", "") for k in
                            ("type", "entry_id", "entry_ids", "expense_id", "project_name", "date", "pdf_filename")}
                    item["description"] = request.form.get(f"item_{i}_description", "").strip()
                    try:
                        item["amount"] = round(float(request.form.get(f"item_{i}_amount", 0)), 2)
                    except ValueError:
                        item["amount"] = 0.0
                    if item["type"] == "time":
                        try:
                            item["hours"] = float(request.form.get(f"item_{i}_hours", 0))
                            item["rate"] = float(request.form.get(f"item_{i}_rate", 0))
                        except ValueError:
                            pass
                    if item["type"] == "manual":
                        try:
                            item["hours"] = float(request.form.get(f"item_{i}_hours", 0)) if request.form.get(f"item_{i}_hours") else None
                            item["rate"] = float(request.form.get(f"item_{i}_rate", 0)) if request.form.get(f"item_{i}_rate") else None
                        except ValueError:
                            pass
                        gst_rate = settings_data.get("gst_rate", 0.05)
                        include_gst = request.form.get(f"item_{i}_gst") == "1"
                        item["item_gst"] = round(item["amount"] * gst_rate, 2) if include_gst else 0.0
                    # Strip empty optional fields
                    item = {k: v for k, v in item.items() if v not in ("", None)}
                    new_line_items.append(item)
                i += 1

            # Add new manual line items
            j = 0
            while f"new_{j}_description" in request.form:
                desc = request.form.get(f"new_{j}_description", "").strip()
                try:
                    hrs_str = request.form.get(f"new_{j}_hours", "").strip()
                    rate_str = request.form.get(f"new_{j}_rate", "").strip()
                    hrs = float(hrs_str) if hrs_str else None
                    rate = float(rate_str) if rate_str else None
                    amt = round(float(request.form.get(f"new_{j}_amount", 0)), 2)
                except ValueError:
                    hrs = rate = None
                    amt = 0.0
                if desc:
                    gst_rate = settings_data.get("gst_rate", 0.05)
                    include_gst = request.form.get(f"new_{j}_gst") == "1"
                    item_gst = round(amt * gst_rate, 2) if include_gst else 0.0
                    item = {"type": "manual", "description": desc, "amount": amt, "item_gst": item_gst}
                    if hrs is not None:
                        item["hours"] = hrs
                    if rate is not None:
                        item["rate"] = rate
                    new_line_items.append(item)
                j += 1

            subtotal = round(sum(li["amount"] for li in new_line_items), 2)
            gst_rate = settings_data.get("gst_rate", 0.05)
            gst = round(sum(li.get("item_gst", li["amount"] * gst_rate if li.get("type") == "time" else 0)
                            for li in new_line_items), 2)

            invoice["line_items"] = new_line_items
            invoice["subtotal"] = subtotal
            invoice["gst"] = gst
            invoice["total"] = round(subtotal + gst, 2)

            storage.save_invoices(inv_list)
            storage.save_time_entries(all_entries)
            storage.save_expenses(all_expenses)
            flash("Invoice updated.", "success")
            return redirect(url_for("invoices"))

        return render_template("invoice_edit.html", invoice=invoice)

    @app.route("/invoices/<inv_id>/status", methods=["POST"])
    def invoices_status(inv_id):
        new_status = request.form.get("status", "").strip()
        if new_status not in ("draft", "sent", "late", "paid"):
            flash("Invalid status.", "error")
            return redirect(url_for("invoices"))
        inv_list = storage.load_invoices()
        for inv in inv_list:
            if inv["id"] == inv_id:
                inv["status"] = new_status
                if new_status == "sent":
                    inv["sent"] = True
                storage.save_invoices(inv_list)
                flash(f"Invoice marked as {new_status}.", "success")
                return redirect(url_for("invoices"))
        flash("Invoice not found.", "error")
        return redirect(url_for("invoices"))

    @app.route("/invoices/<inv_id>/delete", methods=["POST"])
    def invoices_delete(inv_id):
        inv_list = storage.load_invoices()
        invoice = next((i for i in inv_list if i["id"] == inv_id), None)
        if not invoice:
            flash("Invoice not found.", "error")
            return redirect(url_for("invoices"))

        all_entries = storage.load_time_entries()
        all_expenses = storage.load_expenses()

        for item in invoice.get("line_items", []):
            ids_to_uninvoice = set()
            entry_id = item.get("entry_id", "")
            entry_ids_str = item.get("entry_ids", "")
            if entry_id:
                ids_to_uninvoice.add(entry_id)
            if entry_ids_str:
                ids_to_uninvoice.update(entry_ids_str.split(","))
            for e in all_entries:
                if e["id"] in ids_to_uninvoice:
                    e["invoiced"] = False
            expense_id = item.get("expense_id", "")
            for x in all_expenses:
                if x["id"] == expense_id:
                    x["invoiced"] = False

        storage.save_time_entries(all_entries)
        storage.save_expenses(all_expenses)
        inv_list = [i for i in inv_list if i["id"] != inv_id]
        storage.save_invoices(inv_list)

        flash(f"Invoice {invoice.get('invoice_number', '')} deleted.", "success")
        return redirect(url_for("invoices"))

    @app.route("/invoices/<inv_id>/preview")
    def invoices_preview(inv_id):
        inv_list = storage.load_invoices()
        invoice = next((i for i in inv_list if i["id"] == inv_id), None)
        if not invoice:
            flash("Invoice not found.", "error")
            return redirect(url_for("invoices"))
        settings_data = storage.load_settings()
        from timesheet.pdf import _logo_data_uri
        logo_uri = _logo_data_uri(settings_data)
        return render_template("invoice_pdf.html", invoice=invoice,
                               settings=settings_data, logo_uri=logo_uri)

    @app.route("/invoices/<inv_id>/pdf")
    def invoices_pdf(inv_id):
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        inv_list = storage.load_invoices()
        invoice = next((i for i in inv_list if i["id"] == inv_id), None)
        if not invoice:
            flash("Invoice not found.", "error")
            return redirect(url_for("invoices"))
        settings_data = storage.load_settings()
        try:
            pdf_bytes = pdf.generate_invoice_pdf(app, invoice, settings_data, expense_pdf_dir=expense_pdf_dir)
        except Exception as e:
            flash(f"PDF error: {e}", "error")
            return redirect(url_for("invoices"))
        return send_file(BytesIO(pdf_bytes), download_name=f"{invoice['invoice_number']}.pdf",
                         as_attachment=True, mimetype="application/pdf")

    @app.route("/invoices/<inv_id>/send", methods=["POST"])
    def invoices_send(inv_id):
        expense_pdf_dir = storage.DATA_DIR / "expense_pdfs"
        inv_list = storage.load_invoices()
        invoice = next((i for i in inv_list if i["id"] == inv_id), None)
        if not invoice:
            flash("Invoice not found.", "error")
            return redirect(url_for("invoices"))
        recipient = request.form.get("recipient", "").strip()
        if not recipient:
            flash("Recipient email required.", "error")
            return redirect(url_for("invoices"))
        settings_data = storage.load_settings()
        try:
            pdf_bytes = pdf.generate_invoice_pdf(app, invoice, settings_data, expense_pdf_dir=expense_pdf_dir)
        except Exception as e:
            flash(f"PDF error: {e}", "error")
            return redirect(url_for("invoices"))
        from timesheet.mailer import send_invoice_email
        subject = request.form.get("subject", f"{invoice['invoice_number']} from {settings_data.get('business_name','')}")
        body = request.form.get("body", "Please find your invoice attached.")
        try:
            send_invoice_email(settings_data, recipient, subject, body,
                               pdf_bytes, invoice["invoice_number"])
            for i in inv_list:
                if i["id"] == inv_id:
                    i["sent"] = True
                    i["status"] = "sent"
            storage.save_invoices(inv_list)
            flash(f"Invoice sent to {recipient}.", "success")
        except Exception as e:
            flash(f"Failed to send email: {e}", "error")
        return redirect(url_for("invoices"))

    @app.route("/reports/monthly")
    def report_monthly():
        today = dt_date.today()
        try:
            year = int(request.args.get("year", today.year))
            month = int(request.args.get("month", today.month))
            if not (1 <= month <= 12):
                raise ValueError
        except ValueError:
            year, month = today.year, today.month
        month_prefix = f"{year}-{month:02d}"

        projects = storage.load_projects()
        project_map = {p["id"]: p for p in projects}
        entries = [e for e in storage.load_time_entries() if e["date"].startswith(month_prefix)]
        expenses = [x for x in storage.load_expenses() if x["date"].startswith(month_prefix)]

        rows = {}
        for e in entries:
            pid = e["project_id"]
            p = project_map.get(pid, {})
            if pid not in rows:
                rows[pid] = {"name": p.get("name", "Unknown"), "hours": 0, "billable": 0.0, "expenses": 0}
            rows[pid]["hours"] += e["hours"]
            rows[pid]["billable"] += e["hours"] * e.get("rate", p.get("rate", 0))
        for x in expenses:
            pid = x.get("project_id", "")
            p = project_map.get(pid, {})
            row_key = pid if pid else f"__client_{x.get('client_id', '')}"
            if row_key not in rows:
                client_name = ""
                if not pid:
                    clients = storage.load_clients()
                    cm = {c["id"]: c for c in clients}
                    client_name = cm.get(x.get("client_id", ""), {}).get("name", "Unknown")
                rows[row_key] = {"name": p.get("name", client_name or "Unknown"), "hours": 0, "billable": 0.0, "expenses": 0}
            rows[row_key]["expenses"] += x["amount"]

        for r in rows.values():
            r["billable"] = round(r["billable"], 2)
            r["total"] = round(r["billable"] + r["expenses"], 2)

        grand = {
            "hours": sum(r["hours"] for r in rows.values()),
            "billable": round(sum(r["billable"] for r in rows.values()), 2),
            "expenses": round(sum(r["expenses"] for r in rows.values()), 2),
            "total": round(sum(r["total"] for r in rows.values()), 2),
        }

        return render_template("report_monthly.html", rows=rows, grand=grand,
                               year=year, month=month)

    @app.route("/reports/uninvoiced")
    def report_uninvoiced():
        projects = storage.load_projects()
        project_map = {p["id"]: p for p in projects}
        entries = [e for e in storage.load_time_entries() if not e["invoiced"]]
        expenses = [x for x in storage.load_expenses() if not x["invoiced"]]

        rows = {}
        for e in entries:
            pid = e["project_id"]
            p = project_map.get(pid, {})
            if pid not in rows:
                rows[pid] = {"name": p.get("name", "Unknown"),
                             "client_id": p.get("client_id") or "",
                             "hours": 0, "billable": 0.0, "expenses": 0,
                             "entries": [], "expense_items": []}
            entry_amount = round(e["hours"] * e.get("rate", p.get("rate", 0)), 2)
            rows[pid]["hours"] += e["hours"]
            rows[pid]["billable"] += entry_amount
            rows[pid]["entries"].append({**e, "amount": entry_amount})
        for x in expenses:
            pid = x.get("project_id", "")
            row_key = pid if pid else f"__client_{x.get('client_id', '')}"
            p = project_map.get(pid, {})
            if row_key not in rows:
                client_name = ""
                if not pid:
                    clients = storage.load_clients()
                    cm = {c["id"]: c for c in clients}
                    client_name = cm.get(x.get("client_id", ""), {}).get("name", "Unknown")
                rows[row_key] = {"name": p.get("name", client_name or "Unknown"),
                                 "client_id": p.get("client_id") or x.get("client_id", ""),
                                 "hours": 0, "billable": 0.0, "expenses": 0,
                                 "entries": [], "expense_items": []}
            rows[row_key]["expenses"] += x["amount"]
            rows[row_key]["expense_items"].append(x)

        for r in rows.values():
            r["billable"] = round(r["billable"], 2)
            r["total"] = round(r["billable"] + r["expenses"], 2)

        grand = {
            "hours": sum(r["hours"] for r in rows.values()),
            "billable": round(sum(r["billable"] for r in rows.values()), 2),
            "expenses": round(sum(r["expenses"] for r in rows.values()), 2),
            "total": round(sum(r["total"] for r in rows.values()), 2),
        }

        return render_template("report_uninvoiced.html", rows=rows, grand=grand)

    # ── Backup & Export ───────────────────────────────────────────────────────

    @app.route("/backup/download")
    def backup_download():
        import zipfile
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in storage.DATA_DIR.rglob("*"):
                if f.is_file():
                    zf.write(f, f.relative_to(storage.DATA_DIR.parent))
        buf.seek(0)
        filename = f"timesheet-backup-{dt_date.today().isoformat()}.zip"
        return send_file(buf, mimetype="application/zip",
                         as_attachment=True, download_name=filename)

    @app.route("/export/excel")
    def export_excel():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="4F8EF7")

        def style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

        def autofit(ws):
            for col in ws.columns:
                width = max((len(str(c.value or "")) for c in col), default=10) + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 50)

        projects = {p["id"]: p for p in storage.load_projects()}
        clients  = {c["id"]: c for c in storage.load_clients()}

        # ── Time entries ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Time"
        headers = ["Date", "Client", "Project", "Hours", "Rate", "Amount", "Description", "Invoiced"]
        style_header(ws, headers)
        for e in sorted(storage.load_time_entries(), key=lambda x: x["date"]):
            p = projects.get(e.get("project_id", ""), {})
            c = clients.get(p.get("client_id", ""), {})
            rate = e.get("rate", p.get("rate", 0))
            ws.append([
                e["date"], c.get("name", ""), p.get("name", ""),
                e["hours"], rate, round(e["hours"] * rate, 2),
                e.get("description", ""), "Yes" if e.get("invoiced") else "No",
            ])
        autofit(ws)

        # ── Expenses ──────────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Expenses")
        style_header(ws2, ["Date", "Client", "Description", "Amount", "Invoiced", "Receipt"])
        for x in sorted(storage.load_expenses(), key=lambda x: x["date"]):
            c = clients.get(x.get("client_id", ""), {})
            ws2.append([
                x["date"], c.get("name", ""), x.get("description", ""),
                x["amount"], "Yes" if x.get("invoiced") else "No",
                x.get("pdf_filename", ""),
            ])
        autofit(ws2)

        # ── Invoices ─────────────────────────────────────────────────────────
        ws3 = wb.create_sheet("Invoices")
        style_header(ws3, ["Invoice #", "Type", "Client", "Issue Date", "Subtotal", "GST", "Total", "Status"])
        for inv in sorted(storage.load_invoices(), key=lambda x: x.get("invoice_number", "")):
            ws3.append([
                inv.get("invoice_number", ""), inv.get("invoice_type", "time"),
                inv.get("client_name", ""), inv.get("issued_date", ""),
                inv.get("subtotal", 0), inv.get("gst", 0), inv.get("total", 0),
                inv.get("status", "draft"),
            ])
        autofit(ws3)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"timesheet-export-{dt_date.today().isoformat()}.xlsx"
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)

    return app


if __name__ == "__main__":
    app = create_app()
    app.run(debug=True, port=5000)
