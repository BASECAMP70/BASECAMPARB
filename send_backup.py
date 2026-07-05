"""
Weekly timesheet backup emailer.
Generates Excel export, time CSV, and full data ZIP (JSON + PDFs).
Emails all three and copies to Google Drive.
Run via Windows Task Scheduler every Sunday.
"""
import csv
import sys
import zipfile
import smtplib
from io import BytesIO, StringIO
from datetime import date
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from timesheet import storage
from timesheet import pdf as pdf_mod

RECIPIENT    = "scott@basecampinc.ca"
GDRIVE_FOLDER = r"Z:\My Drive\TIMESHEET"

DATA_DIR  = Path(os.path.dirname(os.path.abspath(__file__))) / "timesheet" / "data"
JSON_FILES = [
    "clients.json", "projects.json", "time_entries.json",
    "expenses.json", "invoices.json", "tasks.json",
    "project_tasks.json", "settings.json",
]


def _style_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = __import__("openpyxl").styles.PatternFill("solid", fgColor="1D3557")
        cell.alignment = __import__("openpyxl").styles.Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 50)


def _safe_name(s, max_len=30):
    return "".join(c if c.isalnum() or c in " -_" else "" for c in s).strip()[:max_len]


def build_excel():
    wb = Workbook()
    projects = {p["id"]: p for p in storage.load_projects()}
    clients  = {c["id"]: c for c in storage.load_clients()}

    ws = wb.active
    ws.title = "Time"
    _style_header(ws, ["Date", "Client", "Project", "Hours", "Rate", "Amount", "Description", "Invoiced"])
    all_entries = sorted(storage.load_time_entries(), key=lambda x: x["date"])
    for e in all_entries:
        p    = projects.get(e.get("project_id", ""), {})
        c    = clients.get(p.get("client_id", ""), {})
        rate = e.get("rate", p.get("rate", 0))
        ws.append([
            e["date"], c.get("name", ""), p.get("name", ""),
            e["hours"], rate, round(e["hours"] * rate, 2),
            e.get("description", ""), "Yes" if e.get("invoiced") else "No",
        ])
    _autofit(ws)

    ws2 = wb.create_sheet("Expenses")
    _style_header(ws2, ["Date", "Client", "Description", "Amount", "Invoiced", "Receipt"])
    all_expenses = sorted(storage.load_expenses(), key=lambda x: x["date"])
    for x in all_expenses:
        c = clients.get(x.get("client_id", ""), {})
        ws2.append([
            x["date"], c.get("name", ""), x.get("description", ""),
            x["amount"], "Yes" if x.get("invoiced") else "No",
            x.get("pdf_filename", ""),
        ])
    _autofit(ws2)

    ws3 = wb.create_sheet("Invoices")
    _style_header(ws3, ["Invoice #", "Type", "Client", "Issue Date", "Subtotal", "GST", "Total", "Status"])
    for inv in sorted(storage.load_invoices(), key=lambda x: x.get("invoice_number", "")):
        ws3.append([
            inv.get("invoice_number", ""), inv.get("invoice_type", "time"),
            inv.get("client_name", ""), inv.get("issued_date", ""),
            inv.get("subtotal", 0), inv.get("gst", 0), inv.get("total", 0),
            inv.get("status", "draft"),
        ])
    _autofit(ws3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), all_entries, all_expenses


def build_time_csv(all_entries):
    projects = {p["id"]: p for p in storage.load_projects()}
    clients  = {c["id"]: c for c in storage.load_clients()}

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Client", "Project", "Hours", "Rate", "Amount", "Description", "Invoiced"])
    for e in all_entries:
        p    = projects.get(e.get("project_id", ""), {})
        c    = clients.get(p.get("client_id", ""), {})
        rate = e.get("rate", p.get("rate", 0))
        writer.writerow([
            e["date"], c.get("name", ""), p.get("name", ""),
            e["hours"], rate, round(e["hours"] * rate, 2),
            e.get("description", ""), "Yes" if e.get("invoiced") else "No",
        ])
    return buf.getvalue().encode("utf-8-sig")


def build_data_zip(all_expenses, excel_bytes=None):
    """ZIP containing: JSON data files, expense receipt PDFs, generated invoice PDFs."""
    from timesheet.app import create_app
    app = create_app()
    settings_data = storage.load_settings()
    expense_pdf_dir = DATA_DIR / "expense_pdfs"
    clients = {c["id"]: c for c in storage.load_clients()}

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:

        # Excel export
        if excel_bytes:
            zf.writestr(f"timesheet-backup-{date.today().strftime('%Y-%m-%d')}.xlsx", excel_bytes)

        # JSON data files
        for name in JSON_FILES:
            path = DATA_DIR / name
            if path.exists():
                zf.write(path, f"data/{name}")

        # Expense receipt PDFs — rename from UUID to readable name
        for x in all_expenses:
            fname = x.get("pdf_filename")
            if not fname:
                continue
            pdf_path = expense_pdf_dir / fname
            if not pdf_path.exists():
                continue
            c = clients.get(x.get("client_id", ""), {})
            client_name = _safe_name(c.get("name", "unknown"))
            desc        = _safe_name(x.get("description", "")[:20])
            nice = f"expense-{x['date']}-{client_name}"
            if desc:
                nice += f"-{desc}"
            nice += ".pdf"
            zf.writestr(f"expense_receipts/{nice}", pdf_path.read_bytes())

        # Invoice PDFs — generated on the fly
        for inv in storage.load_invoices():
            try:
                pdf_bytes = pdf_mod.generate_invoice_pdf(
                    app, inv, settings_data, expense_pdf_dir=expense_pdf_dir
                )
                inv_num = inv.get("invoice_number", inv["id"])
                zf.writestr(f"invoices/{inv_num}.pdf", pdf_bytes)
            except Exception as e:
                print(f"  WARNING: could not generate PDF for {inv.get('invoice_number', '?')}: {e}")

    buf.seek(0)
    return buf.getvalue()



def attach(msg, data, filename, mime="application/octet-stream"):
    part = MIMEBase(*mime.split("/", 1))
    part.set_payload(data)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)


def send_backup():
    settings = storage.load_settings()
    smtp_host     = settings.get("smtp_host", "smtp.gmail.com")
    smtp_port     = int(settings.get("smtp_port", 587))
    smtp_user     = settings.get("smtp_user", "")
    smtp_password = settings.get("smtp_password", "")

    if not smtp_user or not smtp_password:
        print("ERROR: SMTP credentials not configured in Settings.")
        sys.exit(1)

    today = date.today().strftime("%Y-%m-%d")

    print("Building Excel export...")
    excel_bytes, all_entries, all_expenses = build_excel()

    print("Building data ZIP (Excel + JSON + PDFs)...")
    zip_bytes = build_data_zip(all_expenses, excel_bytes=excel_bytes)

    subject = f"Timesheet Backup {today}"

    zip_name = f"timesheet-backup-{today}.zip"

    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = RECIPIENT
    msg["Subject"] = subject
    msg.attach(MIMEText(
        f"Weekly timesheet backup attached.\n\nContents of {zip_name}:\n"
        f"  timesheet-backup-{today}.xlsx — Excel export (Time, Expenses, Invoices)\n"
        f"  data/                         — JSON data files\n"
        f"  expense_receipts/             — expense receipt PDFs\n"
        f"  invoices/                     — invoice PDFs",
        "plain"
    ))

    attach(msg, zip_bytes, zip_name)

    print("Sending email...")
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, RECIPIENT, msg.as_string())

    print(f"Backup sent to {RECIPIENT} — subject: {subject}")

    # Copy all three to Google Drive
    folder = Path(GDRIVE_FOLDER)
    for data, name in [(zip_bytes, zip_name)]:
        try:
            (folder / name).write_bytes(data)
            print(f"Copied to Google Drive: {folder / name}")
        except Exception as e:
            print(f"WARNING: Google Drive copy failed for {name}: {e}")


if __name__ == "__main__":
    send_backup()
